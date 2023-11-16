import openpyxl
from random import randint, uniform

# Function to generate random product data
def generate_sample_data():
    product_id = randint(1001, 9999)
    product_name = f"Product_{product_id}"
    price = round(uniform(10.0, 1000.0), 2)
    quantity = randint(1, 50)
    return product_id, product_name, price, quantity

# Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers to the sheet
headers = ["Product ID", "Product Name", "Price", "Quantity"]
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Generate and write 100 sample data rows
for row_num in range(2, 103):  # 2 to 101 (100 rows)
    product_data = generate_sample_data()
    for col_num, data in enumerate(product_data, 1):
        sheet.cell(row=row_num, column=col_num, value=data)

# Save the workbook to a file
workbook.save("sales.xlsx")

print("Sample data has been generated and written to sales.xlsx.")
